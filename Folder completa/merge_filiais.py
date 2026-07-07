import pandas as pd
import openpyxl
import os
import glob
import sys
import datetime
import logging
import traceback
from openpyxl.styles import Border, Side

# Setup logging for merge_filiais
logger = logging.getLogger('merge_filiais')

# Define standard thin border
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def get_master_file(SAVE_FOLDER):
    # Convert V: to C: if needed (V: is mapped to C:)
    actual_folder = SAVE_FOLDER.replace("V:", "C:")
    logger.info(f"Looking for master file in: {actual_folder}")

    matches = glob.glob(os.path.join(actual_folder, "RELA*.xlsx"))
    if matches:
        logger.info(f"Found master file: {matches[0]}")
        print(f"  ✓ Master file found: {matches[0]}")
        return matches[0]

    logger.error(f"Master file not found in {actual_folder}")
    print(f"  ✗ Master file not found in {actual_folder}")
    sys.exit(1)

def get_latest_file(SAVE_FOLDER, prefix):
    today = datetime.datetime.now().strftime("%Y%m%d")

    # Convert V: to C: if needed (V: is mapped to C:)
    actual_folder = SAVE_FOLDER.replace("V:", "C:")

    logger.info(f"Looking for {prefix} file from today ({today})")
    logger.info(f"Search folder: {actual_folder}")

    if not os.path.exists(actual_folder):
        logger.error(f"Folder does not exist: {actual_folder}")
        print(f"  ✗ Folder not found: {actual_folder}")
        sys.exit(1)

    pattern = os.path.join(actual_folder, f"{prefix}_{today}_*.xls")
    logger.info(f"Pattern: {pattern}")
    files = glob.glob(pattern)

    if files:
        latest = max(files, key=os.path.getmtime)
        logger.info(f"Found: {latest}")
        print(f"  ✓ Found: {os.path.basename(latest)}")
        return latest

    # If not found, show what files do exist
    logger.error(f"No file found for {prefix} on {today}")
    print(f"  ✗ No file found for {prefix} on {today}")

    logger.info(f"Files in {actual_folder}:")
    try:
        all_files = glob.glob(os.path.join(actual_folder, f"{prefix}_*.xls"))
        for f in all_files:
            logger.info(f"  - {os.path.basename(f)}")
            print(f"  - {os.path.basename(f)}")
    except Exception as e:
        logger.error(f"Error listing files: {e}")

    sys.exit(1)

# ============================================================
# RUN F — Filiais 1+2+5, Qtd. Pedida+Reservada (combined or individual)
# ============================================================
def run(SAVE_FOLDER, filial_prefix="TODAS FILIAIS"):
    """Merge filial data. filial_prefix can be 'TODAS FILIAIS', 'QTD PEDIDA_RESERVADA filial1', etc."""
    try:
        logger.info(f"Starting merge for SAVE_FOLDER: {SAVE_FOLDER}")
        logger.info(f"Looking for prefix: {filial_prefix}")
        print(f"\n[MERGE] Starting merge process...")

        MASTER_FILE = get_master_file(SAVE_FOLDER)
        logger.info(f"Master file: {MASTER_FILE}")

        wb = openpyxl.load_workbook(MASTER_FILE)
        logger.info(f"Workbook loaded. Sheets: {wb.sheetnames}")

        # --- Filial 1+2+5 sheet ---
        SHEET_NAME = "Filial 1+2+5"
        if SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
            print(f"  ✗ Sheet '{SHEET_NAME}' not found")
            print(f"  Available sheets: {wb.sheetnames}")
            sys.exit(1)

        print(f"Finding exported file with prefix: {filial_prefix}...")
        logger.info(f"Looking for export with prefix: {filial_prefix}")
        filiais_file = get_latest_file(SAVE_FOLDER, filial_prefix)
        logger.info(f"Found file: {filiais_file}")

        print("Reading exported file...")
        df_combined = pd.read_excel(filiais_file)
        logger.info(f"Read {len(df_combined)} rows from export")
        print(f"  ✓ Read {len(df_combined)} rows total")

        ws = wb[SHEET_NAME]
        print("Clearing previous data from Filial 1+2+5...")
        logger.info("Clearing sheet data...")
        ws.delete_rows(4, ws.max_row)

        print("Writing Filial 1+2+5 data...")
        logger.info("Writing Filial 1+2+5 data...")
        for row_idx, row in enumerate(df_combined.itertuples(index=False), start=4):
            for col_idx, value in enumerate(row, start=2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

        # --- Qtd. Pedida+Reservada sheet ---
        QTD_SHEET = "Qtd. Pedida+Reservada"
        if QTD_SHEET not in wb.sheetnames:
            logger.error(f"Sheet '{QTD_SHEET}' not found. Available: {wb.sheetnames}")
            print(f"  ✗ Sheet '{QTD_SHEET}' not found")
            print(f"  Available sheets: {wb.sheetnames}")
            sys.exit(1)

        ws2 = wb[QTD_SHEET]
        print("Clearing previous data from Qtd. Pedida+Reservada...")
        logger.info("Clearing Qtd. Pedida+Reservada data...")
        ws2.delete_rows(5, ws2.max_row)

        df_pivot = df_combined.groupby('CODPROD', as_index=False)['QTPENDENTE'].sum()
        df_pivot = df_pivot.sort_values('CODPROD')

        print("Writing Qtd. Pedida+Reservada data...")
        logger.info("Writing Qtd. Pedida+Reservada data...")
        for row_idx, row in enumerate(df_pivot.itertuples(index=False), start=5):
            ws2.cell(row=row_idx, column=3, value=row.CODPROD).border = border
            ws2.cell(row=row_idx, column=6, value=row.QTPENDENTE).border = border

        logger.info(f"Saving workbook to {MASTER_FILE}...")
        wb.save(MASTER_FILE)
        logger.info("✅ Merge completed successfully")
        print(f"  ✓ Filial 1+2+5 updated with {len(df_combined)} rows")
        print(f"  ✓ Qtd. Pedida+Reservada updated with {len(df_pivot)} rows")
        print(f"\n✅ Done! Master file updated: {MASTER_FILE}")

    except Exception as e:
        logger.error(f"ERROR in run(): {str(e)}\n{traceback.format_exc()}")
        print(f"\n❌ ERROR during merge: {str(e)}")
        print(traceback.format_exc())
        sys.exit(1)

# ============================================================
# RUN BP — Base Produtos
# ============================================================
def run_bp(SAVE_FOLDER):
    try:
        logger.info(f"Starting BP merge for SAVE_FOLDER: {SAVE_FOLDER}")
        print(f"\n[MERGE] Starting Base Produtos merge...")

        MASTER_FILE = get_master_file(SAVE_FOLDER)
        logger.info(f"Master file: {MASTER_FILE}")

        wb = openpyxl.load_workbook(MASTER_FILE)
        logger.info(f"Workbook loaded. Sheets: {wb.sheetnames}")

        BASE_SHEET = "Base Produtos"
        if BASE_SHEET not in wb.sheetnames:
            logger.error(f"Sheet '{BASE_SHEET}' not found. Available: {wb.sheetnames}")
            print(f"  ✗ Sheet '{BASE_SHEET}' not found")
            print(f"  Available sheets: {wb.sheetnames}")
            sys.exit(1)

        print("Finding Base Produtos file...")
        logger.info("Looking for BASE PRODUTOS export...")
        base_file = get_latest_file(SAVE_FOLDER, "BASE PRODUTOS")
        logger.info(f"Found file: {base_file}")

        df_base = pd.read_excel(base_file)
        logger.info(f"Read {len(df_base)} rows from BASE PRODUTOS")

        ws4 = wb[BASE_SHEET]
        # Clear only data columns B-E from row 11 onwards, keep F/G/H formulas intact
        print("Clearing previous data from Base Produtos...")
        logger.info("Clearing Base Produtos data...")
        for row in ws4.iter_rows(min_row=11, max_row=ws4.max_row, min_col=2, max_col=5):
            for cell in row:
                cell.value = None

        print("Writing Base Produtos data...")
        logger.info("Writing Base Produtos data...")
        for row_idx, row in enumerate(df_base.itertuples(index=False), start=11):
            ws4.cell(row=row_idx, column=2, value=row.CODPROD).border = border
            ws4.cell(row=row_idx, column=3, value=row.DESCRICAO).border = border
            ws4.cell(row=row_idx, column=4, value=row.CUSTOREP).border = border
            ws4.cell(row=row_idx, column=5, value=row.DTULTALTCUSTOREP).border = border

        ws4.column_dimensions['E'].width = 18

        logger.info(f"Saving workbook to {MASTER_FILE}...")
        wb.save(MASTER_FILE)
        logger.info("[DONE] BP merge completed successfully")
        print(f"  [DONE] Base Produtos updated with {len(df_base)} rows")
        print(f"\n[DONE] Master file updated: {MASTER_FILE}")

    except Exception as e:
        logger.error(f"ERROR in run_bp(): {str(e)}\n{traceback.format_exc()}")
        print(f"\n[ERROR] ERROR during BP merge: {str(e)}")
        print(traceback.format_exc())
        sys.exit(1)

# ============================================================
# RUN PEDIDOS — Pedidos Compra
# ============================================================
def run_pedidos(SAVE_FOLDER):
    try:
        logger.info(f"Starting PEDIDOS merge for SAVE_FOLDER: {SAVE_FOLDER}")
        print(f"\n[MERGE] Starting Pedidos Compra merge...")

        MASTER_FILE = get_master_file(SAVE_FOLDER)
        logger.info(f"Master file: {MASTER_FILE}")

        wb = openpyxl.load_workbook(MASTER_FILE)
        logger.info(f"Workbook loaded. Sheets: {wb.sheetnames}")

        PEDIDOS_SHEET = "Pedidos Compra"
        if PEDIDOS_SHEET not in wb.sheetnames:
            logger.error(f"Sheet '{PEDIDOS_SHEET}' not found. Available: {wb.sheetnames}")
            print(f"  ✗ Sheet '{PEDIDOS_SHEET}' not found")
            print(f"  Available sheets: {wb.sheetnames}")
            sys.exit(1)

        print("Finding Pedidos Compra file...")
        logger.info("Looking for PEDIDO COMPRAS export...")
        pedidos_file = get_latest_file(SAVE_FOLDER, "PEDIDO COMPRAS")
        logger.info(f"Found file: {pedidos_file}")

        df_pedidos = pd.read_excel(pedidos_file)
        logger.info(f"Read {len(df_pedidos)} rows from PEDIDO COMPRAS")

        ws3 = wb[PEDIDOS_SHEET]
        print("Clearing previous data from Pedidos Compra...")
        logger.info("Clearing Pedidos Compra data...")
        ws3.delete_rows(5, ws3.max_row)

        print("Writing Pedidos Compra data...")
        logger.info("Writing Pedidos Compra data...")
        for row_idx, row in enumerate(df_pedidos.itertuples(index=False), start=5):
            ws3.cell(row=row_idx, column=2, value=row.CODPROD).border = border
            ws3.cell(row=row_idx, column=3, value=row.DESCRICAO).border = border
            ws3.cell(row=row_idx, column=4, value=row.QTPEDIDA).border = border
            ws3.cell(row=row_idx, column=5, value=row.CODFILIAL).border = border
            ws3.cell(row=row_idx, column=6, value=row.DTULTPEDCOMPRA).border = border
            ws3.cell(row=row_idx, column=7, value=row.DTULTALTCOM).border = border

        logger.info(f"Saving workbook to {MASTER_FILE}...")
        wb.save(MASTER_FILE)
        logger.info("[DONE] PEDIDOS merge completed successfully")
        print(f"  [DONE] Pedidos Compra updated with {len(df_pedidos)} rows")
        print(f"\n[DONE] Master file updated: {MASTER_FILE}")

    except Exception as e:
        logger.error(f"ERROR in run_pedidos(): {str(e)}\n{traceback.format_exc()}")
        print(f"\n[ERROR] ERROR during PEDIDOS merge: {str(e)}")
        print(traceback.format_exc())
        sys.exit(1)