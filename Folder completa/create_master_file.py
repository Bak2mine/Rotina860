import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

def create_master_file(filepath):
    """Create the master RELAÇÃO PRODUTOS.xlsx file with all required sheets."""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)

    # ============================================================
    # SHEET 1: Base Produtos
    # ============================================================
    ws1 = wb.create_sheet("Base Produtos")
    headers1 = ["", "CODPROD", "DESCRICAO", "CUSTOREP", "DTULTALTCUSTOREP", "", "", ""]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=10, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 40
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 18

    # ============================================================
    # SHEET 2: Pedidos Compra
    # ============================================================
    ws2 = wb.create_sheet("Pedidos Compra")
    headers2 = ["", "CODPROD", "DESCRICAO", "QTPEDIDA", "CODFILIAL", "DTULTPEDCOMPRA", "DTULTALTCOM"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws2.column_dimensions['A'].width = 3
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15

    # ============================================================
    # SHEET 3: Filial 1+2+5
    # ============================================================
    ws3 = wb.create_sheet("Filial 1+2+5")
    headers3 = ["", "CODPROD", "DESCRICAO", "CODFILIAL", "QTPENDENTE"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws3.column_dimensions['A'].width = 3
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 40
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 15

    # ============================================================
    # SHEET 4: Qtd. Pedida+Reservada
    # ============================================================
    ws4 = wb.create_sheet("Qtd. Pedida+Reservada")
    headers4 = ["", "", "CODPROD", "", "", "QTPENDENTE"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    ws4.column_dimensions['A'].width = 3
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['F'].width = 15

    # Save the workbook
    wb.save(filepath)
    print(f"  ✓ Master file created: {filepath}")
    return filepath


if __name__ == "__main__":
    # Create in current directory
    filepath = "RELAÇÃO PRODUTOS.xlsx"
    create_master_file(filepath)
